"""
excel_engine.py
---------------

FinPilot AI

Handles ALL Excel operations.

Responsibilities:
- Open workbook
- Save workbook
- Read transactions
- Write transactions (with normalization)
- Find next empty row
"""

from datetime import datetime
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ----------------------------------------------------
# CONFIG
# ----------------------------------------------------

EXCEL_FILE = "data/Budget Tracker Final...xlsx"

TRANSACTION_SHEET = "Transactions"

VALID_TYPES = {"Income", "Expense", "Savings", "Debt"}


# ----------------------------------------------------
# WORKBOOK
# ----------------------------------------------------

def open_workbook() -> Workbook:
    """Opens the Excel workbook. Raises FileNotFoundError with a clear
    message if the file is missing, so callers can show a friendly error."""

    try:
        return load_workbook(EXCEL_FILE)
    except FileNotFoundError as exc:
        raise FileNotFoundError(
            f"Excel file not found at '{EXCEL_FILE}'. "
            "Check that the data folder wasn't moved or renamed."
        ) from exc


# ----------------------------------------------------
# TRANSACTIONS
# ----------------------------------------------------

def get_transaction_sheet(workbook: Workbook) -> Worksheet:
    """Returns the Transactions worksheet."""

    if TRANSACTION_SHEET not in workbook.sheetnames:
        raise KeyError(
            f"'{TRANSACTION_SHEET}' sheet is missing from the workbook."
        )

    return workbook[TRANSACTION_SHEET]


# ----------------------------------------------------
# NEXT EMPTY ROW
# ----------------------------------------------------

def get_next_empty_row(sheet: Worksheet) -> int:

    row = 4  # row 1 = column letters, row 2 = separator, row 3 = header

    while sheet.cell(row=row, column=1).value is not None:
        row += 1

    return row


# ----------------------------------------------------
# NORMALIZATION
# ----------------------------------------------------

def normalize_transaction(transaction: dict[str, Any]) -> dict[str, Any]:
    """
    Fills in anything the AI left vague ("today", "Unknown", missing
    month/year) with real values, and validates the type field.

    This is the single place that guarantees every row written to
    Excel has a real date, month and year — no more garbage rows.

    IMPORTANT: month/year are derived from the ACTUAL transaction date
    (whatever date ends up in clean["date"]), never from "today" in
    isolation. Previously month/year defaulted to the real-world
    current month whenever they weren't explicitly supplied, even if
    a specific past/future date was given — which caused every
    transaction to be tagged with whatever month it happened to be
    "right now", regardless of the date entered. That made the
    per-month dashboard views appear frozen/not updating, since all
    data collapsed into a single month bucket.
    """

    from datetime import timedelta

    today = datetime.now()
    clean = dict(transaction)

    # ---- date (resolve to a real datetime first) ----
    raw_date = str(clean.get("date", "")).strip().lower()

    if raw_date in ("", "today", "none"):
        resolved_date = today
    elif raw_date == "yesterday":
        resolved_date = today - timedelta(days=1)
    else:
        resolved_date = None
        original_date_str = str(transaction.get("date", "")).strip()
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d %B %Y", "%d %b %Y"):
            try:
                resolved_date = datetime.strptime(original_date_str, fmt)
                break
            except (TypeError, ValueError):
                continue
        # If the date string couldn't be parsed, fall back to today
        # rather than silently writing an un-derivable month/year.
        if resolved_date is None:
            resolved_date = today

    clean["date"] = resolved_date.strftime("%d-%m-%Y")

    # ---- month (derived from the resolved date, not "today") ----
    raw_month = str(clean.get("month", "")).strip().lower()
    if raw_month in ("", "none", "unknown", "not specified"):
        clean["month"] = resolved_date.strftime("%B")
    else:
        clean["month"] = transaction.get("month")

    # ---- year (derived from the resolved date, not "today") ----
    raw_year = clean.get("year")
    if not raw_year or str(raw_year).strip().lower() in ("none", "unknown", "0"):
        clean["year"] = resolved_date.year
    else:
        try:
            clean["year"] = int(raw_year)
        except (TypeError, ValueError):
            clean["year"] = resolved_date.year

    # ---- type ----
    t_type = str(clean.get("type", "")).strip()
    if t_type not in VALID_TYPES:
        clean["type"] = "Expense"
    else:
        clean["type"] = t_type

    # ---- amount ----
    try:
        clean["amount"] = float(clean.get("amount", 0) or 0)
    except (TypeError, ValueError):
        clean["amount"] = 0.0

    # ---- merchant / category / description ----
    clean["merchant"] = clean.get("merchant") or "Unspecified"
    clean["category"] = clean.get("category") or "Other Expenses"
    clean["description"] = clean.get("description") or ""

    return clean


# ----------------------------------------------------
# APPEND TRANSACTION
# ----------------------------------------------------

def append_transaction(sheet: Worksheet, transaction: dict[str, Any]) -> int:
    """Writes a normalized transaction into the next empty row."""

    clean = normalize_transaction(transaction)

    row = get_next_empty_row(sheet)

    sheet.cell(row=row, column=1).value = clean["date"]
    sheet.cell(row=row, column=2).value = clean["type"]
    sheet.cell(row=row, column=3).value = clean["category"]
    sheet.cell(row=row, column=4).value = clean["merchant"]
    sheet.cell(row=row, column=5).value = clean["amount"]
    sheet.cell(row=row, column=6).value = clean["description"]
    sheet.cell(row=row, column=7).value = 0.99
    sheet.cell(row=row, column=8).value = clean["month"]
    sheet.cell(row=row, column=9).value = clean["year"]

    return row


# ----------------------------------------------------
# SAVE TRANSACTION
# ----------------------------------------------------

def write_transaction(transaction: dict[str, Any]) -> int:
    """Saves one (normalized) transaction to Excel."""

    workbook = open_workbook()

    try:
        sheet = get_transaction_sheet(workbook)
        row = append_transaction(sheet, transaction)
        workbook.save(EXCEL_FILE)
    finally:
        workbook.close()

    return row


# ----------------------------------------------------
# READ ALL TRANSACTIONS
# ----------------------------------------------------

def read_transactions() -> list[tuple]:

    workbook = load_workbook(EXCEL_FILE, data_only=True)

    try:
        sheet = get_transaction_sheet(workbook)

        data = []

        for row in sheet.iter_rows(min_row=4, values_only=True):
            if row[0] is not None:
                data.append(row)

        return data
    finally:
        workbook.close()


# ----------------------------------------------------
# SEARCH
# ----------------------------------------------------

def search_transactions(keyword: str) -> list[tuple]:

    keyword = keyword.lower()

    results = []

    for row in read_transactions():
        text = " ".join(str(i) for i in row if i is not None).lower()
        if keyword in text:
            results.append(row)

    return results


# ----------------------------------------------------
# TOTALS
# ----------------------------------------------------

def get_total_income() -> float:
    return sum(row[4] for row in read_transactions() if row[1] == "Income")


def get_total_expense() -> float:
    return sum(row[4] for row in read_transactions() if row[1] == "Expense")


# ----------------------------------------------------
# TEST
# ----------------------------------------------------

if __name__ == "__main__":

    print("=" * 60)
    print("FinPilot Excel Engine")
    print("=" * 60)
    print()
    print("Transactions:")

    transactions = read_transactions()
    print(len(transactions), "transactions found")
    print()
    print("Income :", get_total_income())
    print("Expense:", get_total_expense())
    print()
    print("Search Domino")
    print(search_transactions("Domino"))