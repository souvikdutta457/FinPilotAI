"""
dashboard.py
------------
FinPilot AI

Reads dashboard information and calculates dynamic totals purely
from the Transactions sheet, driven entirely by the month passed
in from the GUI.

Note: the per-month sheets (January..December) contain a static
"Budget" column (H16:H21) which IS legitimately read per sheet.
Their D8:D11 "baseline" cells are NOT used here - they are
identical placeholder values across every month in the workbook
(a leftover from cloning the template), so Income/Expense/Debt/
Savings are computed live from the Transactions sheet instead.

DEBUG: temporary trace prints are included below, guarded by the
DEBUG flag. Set DEBUG = False (or delete the print lines) once
you've confirmed everything is working as expected.
"""

from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from . import excel_engine

# ---------------- CONFIG ---------------- #

EXCEL_FILE = "data/Budget Tracker Final...xlsx"

MONTH_SHEETS = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}

# Set to False to silence the temporary debug trace prints.
DEBUG = True


def _current_month_sheet() -> str:
    return MONTH_SHEETS[datetime.now().month]


def _transactions_for_month(target_month: str) -> list[tuple]:
    """
    Returns every transaction row whose stored month (column index 7)
    matches the given month name, case-insensitively, with whitespace
    stripped on both sides. No sheet is ever used as a fallback, so
    results are always genuinely driven by the selected month and
    never leak another month's data.
    """

    all_txns = excel_engine.read_transactions()
    target_clean = target_month.strip().lower()

    matched = [
        t for t in all_txns
        if len(t) > 7 and t[7] is not None
        and str(t[7]).strip().lower() == target_clean
    ]

    if DEBUG:
        distinct_months = sorted({
            str(t[7]).strip() if len(t) > 7 and t[7] is not None else "<blank>"
            for t in all_txns
        })
        print(f"[dashboard.py DEBUG] requested month = {target_month!r}")
        print(f"[dashboard.py DEBUG] total rows in sheet = {len(all_txns)}")
        print(f"[dashboard.py DEBUG] distinct month values in sheet = {distinct_months}")
        print(f"[dashboard.py DEBUG] rows matched for {target_month!r} = {len(matched)}")

    return matched


# ---------------- DASHBOARD ---------------- #


def get_dashboard_data(month_sheet: str | None = None) -> dict[str, float]:
    """
    Calculates Income / Expense / Debt / Savings / Balance for the
    given month, purely from the Transactions sheet.

    A month with no transactions correctly returns zeros - it never
    falls back to another sheet's cached/placeholder values, so the
    dashboard always reflects the month actually selected in the GUI.
    """

    target_month = month_sheet or _current_month_sheet()

    month_txns = _transactions_for_month(target_month)

    income = sum(float(t[4] or 0) for t in month_txns if t[1] == "Income")
    expense = sum(float(t[4] or 0) for t in month_txns if t[1] == "Expense")
    debt = sum(float(t[4] or 0) for t in month_txns if t[1] == "Debt")
    savings = sum(float(t[4] or 0) for t in month_txns if t[1] == "Savings")

    result = {
        "income": income,
        "expense": expense,
        "debt": debt,
        "savings": savings,
        "balance": income - expense - debt,
    }

    if DEBUG:
        print(f"[dashboard.py DEBUG] get_dashboard_data({target_month!r}) -> {result}")

    return result


# ---------------- EXPENSE BREAKDOWN ---------------- #


def get_expense_breakdown(month_sheet: str | None = None) -> dict[str, dict[str, float]]:
    """
    Returns budget-vs-actual per category for the given month.

    Budget targets are read from that specific month's sheet (H16:H21)
    - these ARE legitimately per-sheet values and are read dynamically
    from whichever sheet matches the selected month, never hardcoded.

    Actual spend is calculated live from the Transactions sheet,
    filtered to that month only.
    """

    target_month = month_sheet or _current_month_sheet()

    rows = {
        "Bus": 16, "Uber/Rapido": 17, "Dining Out": 18,
        "Other Expenses": 19, "Food Outside": 20, "Food Order": 21,
    }

    expenses: dict[str, dict[str, float]] = {
        category: {"budget": 0.0, "actual": 0.0} for category in rows
    }

    workbook = load_workbook(EXCEL_FILE, data_only=True)
    try:
        if target_month in workbook.sheetnames:
            sheet = workbook[target_month]
            for category, row in rows.items():
                expenses[category]["budget"] = sheet[f"H{row}"].value or 0
        elif DEBUG:
            print(f"[dashboard.py DEBUG] sheet {target_month!r} not found in workbook.sheetnames")
    finally:
        workbook.close()

    month_txns = _transactions_for_month(target_month)

    for t in month_txns:
        t_type = t[1]
        t_category = t[2]
        t_amount = float(t[4] or 0)

        if t_type != "Expense":
            continue

        if t_category in expenses:
            expenses[t_category]["actual"] += t_amount
        else:
            # Catch-all for custom/unlisted categories
            expenses["Other Expenses"]["actual"] += t_amount

    if DEBUG:
        print(f"[dashboard.py DEBUG] get_expense_breakdown({target_month!r}) -> {expenses}")

    return expenses


# ---------------- AI INSIGHTS ---------------- #


def generate_ai_insights(month_sheet: str | None = None) -> list[str]:
    """Generates spending insights for the given month's Budget vs Actual."""

    expenses = get_expense_breakdown(month_sheet)

    insights = []

    for category, values in expenses.items():
        budget = values["budget"]
        actual = values["actual"]

        if actual > budget:
            insights.append(f"⚠ {category} exceeded budget by ₹{actual - budget:,.2f}")
        elif actual < budget:
            insights.append(f"✅ {category} is under budget by ₹{budget - actual:,.2f}")
        else:
            insights.append(f"✔ {category} is exactly on budget.")

    return insights


# ---------------- TEST ---------------- #

if __name__ == "__main__":

    for month in ("July", "January", "August"):
        print(f"--- {month} ---")
        print("Dashboard:", get_dashboard_data(month))
        print("Breakdown:", get_expense_breakdown(month))
        print()