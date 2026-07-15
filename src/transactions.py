"""
transactions.py
----------------
FinPilot AI

Handles reading transactions from the Excel workbook.
"""

from typing import Any

from openpyxl import load_workbook

EXCEL_FILE = "data/Budget Tracker Final...xlsx"

# Row 1 = column letters (A, B, C...), row 2 = dashed separator,
# row 3 = real header, row 4 = first data row.
FIRST_DATA_ROW = 4


def _load_transaction_rows() -> list[tuple]:
    """Shared helper: opens the workbook once and returns raw rows."""

    workbook = load_workbook(EXCEL_FILE, data_only=True)

    try:
        sheet = workbook["Transactions"]

        rows = []
        for row in sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
            if row[0] is not None:
                rows.append(row)

        return rows
    finally:
        workbook.close()


def get_recent_transactions(limit: int = 10) -> list[dict[str, Any]]:
    """
    Returns the most recent transactions from the Transactions sheet,
    latest first, as dicts.

    Args:
        limit: Number of recent transactions to return.
    """

    rows = _load_transaction_rows()

    transactions = [
        {
            "date": row[0],
            "type": row[1],
            "category": row[2],
            "merchant": row[3],
            "amount": row[4],
            "description": row[5] if len(row) > 5 else "",
        }
        for row in rows
    ]

    # Latest transaction first
    transactions.reverse()

    return transactions[:limit]


def get_all_transactions() -> list[tuple]:
    """
    Returns every transaction in the workbook as raw tuples.
    Useful for charts, analytics and search.
    """

    return _load_transaction_rows()


# ---------------- TEST ---------------- #

if __name__ == "__main__":

    print("=" * 70)
    print("FinPilot AI - Recent Transactions")
    print("=" * 70)

    for t in get_recent_transactions():
        print(f"{t['date']} | {t['type']} | {t['category']} | "
              f"{t['merchant']} | ₹{t['amount']}")
